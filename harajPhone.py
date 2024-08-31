from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time
import random
import sys
import openpyxl
from openpyxl import load_workbook
import pandas as pd
# =============================== LenksHaraj =================================================
try:
    f = open("LenksHaraj.txt", encoding='utf-8' )
    urls=f.read().split('<|>')
    f.close()
except:
    print('\n'*5,'Make sure that the "LenksHaraj.txt" file is written correctly')
    time.sleep(8)
    sys.exit()
# =============================== config =================================================
try:
    f = open("config.txt", encoding='utf-8' )
    ads_number=int(f.readline().split('The number of ads to restart the window :')[1].strip())
    times=float(f.readline().split('TIME(s) :')[1].strip())
    f.close()

except:
    print('\n'*5,'Make sure that the .confg.txt file is written correctly')
    time.sleep(8)
    sys.exit()
# ====================================== create xlsx ==========================================
def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames[0]

colum = {'Phone':[],'title':[],'city':[],'user':[]}
data=pd.DataFrame(colum)  
filename=str(random.randint(0,99999))+'.xlsx'
data.to_excel(filename,sheet_name='sheet1',index=False)    
# ====================================== loop ==========================================
service = Service(executable_path='./chromedriver.exe')
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)

for t in urls:
    if (urls.index(t)+1)%ads_number==0:
        driver.quit()
        time.sleep(2)
        service = Service(executable_path='./chromedriver.exe')
        options = webdriver.ChromeOptions()
        driver = webdriver.Chrome(service=service, options=options)
    try:
        driver.get(t)
        title=driver.find_element(by=By.XPATH,value="//*[@id='__next']/div[2]/div[2]/div[1]/div[2]/div[1]/div/h1").text
        city=driver.find_element(by=By.XPATH,value='//*[@id="__next"]/div[2]/div[2]/div[1]/div[2]/div[1]/div/div/div[1]/span[1]/a/span').text
        user=driver.find_element(by=By.XPATH,value='//*[@id="__next"]/div[2]/div[2]/div[1]/div[2]/div[1]/div/div/div[2]/span/div[2]/a').text
        driver.find_element(by=By.XPATH,value='//*[@id="__next"]/div[2]/div[2]/div[1]/div[2]/div[2]/div/button').click()
        time.sleep(1.5)
        phone=driver.find_element(by=By.XPATH,value='/html/body/reach-portal[2]/div/div[2]/div/div/div/a[2]/div[2]').text
        print(title ,'-',city,'-',user,'-',phone)
        # ====================================== create len ======================================
        new_row=[phone,title,city,user]
        wbl = openpyxl.load_workbook(filename=filename)
        sheet = wbl[get_sheetnames_xlsx(filename)]
        sheet.append(new_row)
        wbl.save(filename)
        # ========================================================================================
        time.sleep(times)
    except:
        continue
driver.quit()